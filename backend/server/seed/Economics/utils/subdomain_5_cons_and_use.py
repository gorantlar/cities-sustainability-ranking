# This code gets the cities list per state from www.city-data.com
# And store it in data/ state-wise
import time

import requests
import xlsxwriter
from bs4 import BeautifulSoup
import csv
import openpyxl
import os
import glob
import difflib

def get_all_summed(section):
    dollars = 0
    # for all li
    lists = list(section.find_all("li"))
    for li in lists:
        val = li.text
        n = len(val)
        # print(val)

        left_par = 0
        while left_par < n:
            while left_par < n and val[left_par] != "(":
                left_par += 1

            right_par = left_par + 1
            while right_par < n and val[right_par] != ")":
                right_par += 1

            # print("$",dollars,": "+val[left_par + 2: right_par])
            # + 2 for ($

            temp = val[left_par + 2: right_par]
            if temp:
                try:
                    dollars += float(temp)
                except ValueError:
                    print("ValueError: ",temp)
            left_par = right_par + 1
    return dollars

def get_cons_and_use_data(link):
    print(link)
    # link = "https://www.city-data.com/city/Indianapolis-Indiana.html"
    html_text = requests.get(link).text
    soup = BeautifulSoup(html_text, 'html.parser')
    cityDetails = ["N" for _ in consumption_and_use]

    # id = "households-by-units"
    section = soup.find("section", id = "government-finances")
    if(not section):
        return False, cityDetails

    # id = "govFinancesE"
    govFinancesE = section.find("div", id = "govFinancesE")
    if govFinancesE:
        cityDetails[0] = get_all_summed(govFinancesE)

    # id = "govFinancesR"
    govFinancesR = section.find("div", id = "govFinancesR")
    if govFinancesR:
        cityDetails[1] = get_all_summed(govFinancesR)

    # id = "govFinancesD"
    govFinancesD = section.find("div", id = "govFinancesD")
    if govFinancesD:
        cityDetails[2] = get_all_summed(govFinancesD)

    # id = "govFinancesC"
    govFinancesC = section.find("div", id = "govFinancesC")
    if govFinancesC:
        cityDetails[3] = get_all_summed(govFinancesC)



    keys = list(section.find_all("li", class_ ="list-group-item list-group-item-info"))
    # 1+ for <li class="list-group-item active">Unemployment by race in 2019</li>
    values = list(section.find_all("li", class_ ="list-group-item"))[1:]
    # negative buffer
    valIndex = 0
    for i in range(len(keys)):
        # print(keys[i].text, " male: ", values[i*2].text, "female: ", values[i*2+1].text)
        detailsList = []
        # for headers
        valIndex += 1
        while valIndex < len(values) and \
                "rate" not in values[valIndex].text.lower():
            detailsList.append(values[valIndex].text)
            valIndex += 1

        for det in detailsList:
            details = det.split("%")
            if "females" in details[1].lower():
                cityDetails[keys[i].text + " females"] = details[0]+"%"
            elif "males" in details[1].lower():
                cityDetails[keys[i].text + " males"] = details[0]+"%"

    # print(link, ": ", cityDetails)
    return True, cityDetails



# def getCitiesMaps(path):
#     citiesMap = {}
#     with open(path, "r") as readObj:
#         csv_reader = csv.reader(readObj)
#         header = next(csv_reader)
#
#         for row in csv_reader:
#             key = row[1]+":"+row[4]+":"+row[2]
#             citiesMap[key] = row
#
#     return [header, citiesMap]

def getAllCitiesMap(housesBaseURL):
    path = os.getcwd()
    excel_files = glob.glob(os.path.join("../data/", "*.xlsx"))
    cityDataCities = {}

    for file in excel_files:
        wrkbk = openpyxl.load_workbook(file)
        sh = wrkbk.active
        fileName = file.split("\\")[-1]
        stateName = " ".join(fileName.split()[:-1]).lower()

        # for testing 12 cities per state
        # for row in sh.iter_rows(min_row=0, min_col=0, max_row=12, max_col=3):
        for row in sh.iter_rows(min_row=0, min_col=0):
            cityName = row[0].value
            href = row[2].value
            if cityName and href:
                cityName = cityName.lower().replace("-", " ")
                cityDataCities[cityName + ":" + stateName] = housesBaseURL + href
                # print(cityName+ ":" + stateName)
    return cityDataCities


if __name__ == '__main__':
    #    income_base = "www.city-data.com/income/income-Andover-Florida.html"
    housesBaseURL = "https://www.city-data.com/city/"

    # As details are per races
    consumption_and_use = ["government_finances_expenditure_per_resident_in_2018($)",
    "government_finances_revenue_per_resident_in_2018($)",
    "government_finances_debt_per_resident_in_2018($)",
    "government_finances_cash_and_securities_per_resident_in_2018($)"
    ]


    exceptionsFor500 = {
        "nashville:tennessee": "nashville davidson:tennessee",
        "lexington:kentucky": "lexington fayette:kentucky",
        "san buenaventura:california": 'san buenaventura (ventura):california'
    }

    # states not present on city-data.com
    exceptionsStates = {"puerto rico", "village of islands"}

    # name of csv output file
    outputFile = open("../output/output.csv", 'w', newline='', encoding='utf-8')
    csvWriter = csv.writer(outputFile)

    def writeCityDetails(row, cityIncomeDetails):
        row.extend(cityIncomeDetails)
        print("Writing details for: ", row[:5], )
        csvWriter.writerow(row)

    # For Manual-review
    toCheck = []
    totalPassed = 0
    totalFailed = 0
    targetFailed = 0
    # generate city-data cities map
    cityDataCities = getAllCitiesMap(housesBaseURL)

    # might be used in findClosestMatch
    cityDataCitiesList = cityDataCities.keys()


    def findClosestMatch(key):
        return difflib.get_close_matches(key, cityDataCitiesList, cutoff=0.8, n=3)


    with open("../final500Cities.csv", "r") as readObj:
        csvReader = csv.reader(readObj)
        headers = next(csvReader)
        baseHeadersLen = len(headers)


        # add new columns
        def addNewColumns():
            headers.extend(consumption_and_use)


        addNewColumns()
        # writing the headers
        csvWriter.writerow(headers)
        # print("headers: ", headers)

        for row in csvReader:
            cityName = row[1].lower().replace("-", " ")
            stateName = row[2].lower()
            key = cityName + ":" + stateName
            cityHouseDetails = {}

            if key in cityDataCities:
                status, cityHouseDetails = get_cons_and_use_data(cityDataCities[key])
                if status:
                    totalPassed += 1
                else:
                    totalFailed += 1
                    toCheck.append([key, cityDataCities[key], cityHouseDetails])
                time.sleep(1)
            elif key in exceptionsFor500:
                status, cityHouseDetails = get_cons_and_use_data(cityDataCities[exceptionsFor500[key]])
                if status:
                    totalPassed += 1
                else:
                    totalFailed += 1
                    toCheck.append([key, cityDataCities[exceptionsFor500[key]], cityHouseDetails])
                time.sleep(1)
            else:
                if stateName not in exceptionsStates:
                    # closestMatchList = findClosestMatch(key)
                    # print("closest match for ", key, " are ", closestMatchList)
                    toCheck.append([key, cityHouseDetails])
                    totalFailed += 1
            writeCityDetails(row, cityHouseDetails)
            # break

    # todo
    # Target cities score (check with 500 cities)

    # print("toCheck ", toCheck)
    print("totalCities= ", len(cityDataCities))
    print("totalPassed= ", totalPassed, " and totalFailed= ", totalFailed)
    print("toCheck: ", toCheck)

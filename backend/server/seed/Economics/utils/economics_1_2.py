# This code gets the cities list per state from www.city-data.com
# And store it in data/ state-wise
import time

import requests
from bs4 import BeautifulSoup
import csv
import openpyxl
import os
import glob
import difflib

toCheck = []

def get_all_details(link):
    html_text = requests.get(link).text
    soup = BeautifulSoup(html_text, 'html.parser')
    errors = []

    cityDetails = {}
    # --------------------------------------------------------------------------------------------------------
    value_for = ["col_1"]
    section = soup.find("section", id = "religion")
    try:
        table = section.find("div", class_ = "table-responsive")
        denom = 0
        numer = 0
        religion_list = []
        for part in table.find_all("tr"):
            val = ''
            # religion
            td_ptr = part.find_next("td")
            name = td_ptr.text
            if name not in religion_list:
                religion_list.append(name)
            else:
                continue

            # value_1
            td_ptr = td_ptr.find_next("td")
            print(f'{link}')
            if '-' in td_ptr.text:
                continue
            else:
                val = int(td_ptr.text.replace(",", ""), 10)

            if name.lower().strip() == "other":
                numer = val
            print(val)
            denom += val
        print(numer, " and  ", denom)
        print(f'{link}: {numer/denom}')
            # cityDetails[value_for[0]] = val
        return str(round((numer/denom) * 100)) + '%'
    except AttributeError:
        errors.append([link, "'median-income' section not present"])
    return "N"
    # --------------------------------------------------------------------------------------------------------
    # value_for = ["Median gross rent in 2019"]
    # section = soup.find("section", id = "median-rent")
    # try:
    #     cityDetails[value_for[0]] = section.text.split(":")[1]
    # except AttributeError:
    #     errors.append([link, "'median-rent' section not present"])
    #
    # # --------------------------------------------------------------------------------------------------------
    # value_for = ["Cost of living index in March 2019"]
    # section = soup.find("section", id="cost-of-living-index")
    # try:
    #     cityDetails[value_for[0]] = section.text.split(":")[1].split("(")[0]
    # except AttributeError:
    #     errors.append([link, "'cost-of-living-index' section not present"])
    #
    # # --------------------------------------------------------------------------------------------------------
    # value_for = ["Percentage of residents living in poverty in 2019"]
    # section = soup.find("section", id="poverty-level")
    # try:
    #     cityDetails[value_for[0]] = section.text.split(":")[1].split("(")[0]
    # except AttributeError:
    #     errors.append([link, "'poverty-level' section not present"])
    #
    # # --------------------------------------------------------------------------------------------------------
    # value_for = ["Unemployment in November 2020"]
    # section = soup.find("section", id="unemployment")
    # try:
    #     # city
    #     td_ptr = section.find_next("td")
    #     # value
    #     td_ptr = td_ptr.find_next("td")
    #     val = td_ptr.text
    #     cityDetails[value_for[0]] = val
    #
    # except AttributeError:
    #     errors.append([link, "'unemployment' section not present"])
    #
    # if errors:
    #     toCheck.append(errors)
    # print(f'{link}: {cityDetails}')
    return cityDetails


# def getCitiesMaps(path):
#     citiesMap = {}
#     with open(path, "r") as readObj:
#         csv_reader = csv.reader(readObj)
#         header = next(csv_reader)
#
#         for row in csv_reader:
#             key = row[1]+":"+row[4]+":"+row[2]
#             citiesMap[key] = row
#
#     return [header, citiesMap]

def getAllCitiesMap():
    path = os.getcwd()
    excel_files = glob.glob(os.path.join("data/", "*.xlsx"))
    cityDataCities = {}

    for file in excel_files:
        wrkbk = openpyxl.load_workbook(file)
        sh = wrkbk.active
        fileName = file.split("\\")[-1]
        stateName = " ".join(fileName.split()[:-1]).lower().replace("data/", "")

        # for testing 12 cities per state
        # for row in sh.iter_rows(min_row=0, min_col=0, max_row=12, max_col=3):
        for row in sh.iter_rows(min_row=0, min_col=0, max_col=3):
            cityName = row[0].value
            href = row[2].value
            if cityName and href:
                cityName = cityName.lower().replace("-", " ")
                cityDataCities[cityName + ":" + stateName] = baseURL + href
                # print(cityName+ ":" + stateName)
    return cityDataCities


if __name__ == '__main__':
    # income_base = "www.city-data.com/income/income-Andover-Florida.html"
    baseURL = "https://www.city-data.com/city/"

    # all new columns
    column_headers = ["Religious Diversity(Other non-Christian religious groups)"]

    exceptionsFor500 = {
        "nashville:tennessee": "nashville davidson:tennessee",
        "lexington:kentucky": "lexington fayette:kentucky",
        "san buenaventura:california": 'san buenaventura (ventura):california'
    }

    # states not present on city-data.com
    exceptionsStates = {"puerto rico", "village of islands"}

    # name of csv output file
    outputFile = open("output/output.csv", 'w', newline='', encoding='utf-8')
    csvWriter = csv.writer(outputFile)



    # For Manual-review
    prev_size_of_to_check = 0
    totalPassed = 0
    totalFailed = 0
    targetFailed = 0
    # generate city-data cities map
    cityDataCities = getAllCitiesMap()
    cityDataCitiesList = cityDataCities.keys()
    # print(cityDataCitiesList)

    def findClosestMatch(key):
        return difflib.get_close_matches(key, cityDataCitiesList, cutoff=0.8, n=3)


    with open("../../final500Cities.csv", "r") as readObj:
        csvReader = csv.reader(readObj)
        headers = next(csvReader)
        baseHeadersLen = len(headers)


        # add new columns
        def addNewColumns():
            headers.extend(column_headers)


        addNewColumns()
        # writing the headers
        csvWriter.writerow(headers)
        # print("headers: ", headers)

        for row in csvReader:
            cityName = row[1].lower().replace("-", " ")
            stateName = row[2].lower()
            key = cityName + ":" + stateName
            none_percent = {}

            print("key ", key)
            if key in cityDataCities:
                none_percent = get_all_details(cityDataCities[key])
                current_to_check_length = len(toCheck)

                if prev_size_of_to_check == current_to_check_length:
                    totalPassed += 1
                else:
                    totalFailed += 1

            elif key in exceptionsFor500:
                none_percent = get_all_details(cityDataCities[exceptionsFor500[key]])
                current_to_check_length = len(toCheck)

                if prev_size_of_to_check == current_to_check_length:
                    totalPassed += 1
                else:
                    totalFailed += 1

            else:
                if stateName not in exceptionsStates:
                    # closestMatchList = findClosestMatch(key)
                    # print("closest match for ", key, " are ", closestMatchList)
                    toCheck.append(key)
                    totalFailed += 1

            time.sleep(3)
            prev_size_of_to_check = len(toCheck)
            row.append(none_percent)
            csvWriter.writerow(row)
            #break
    # todo
    # Target cities score (check with 500 cities)


    # error logging
    errors = open("error_log.txt", "w")
    for row in toCheck:
        errors.write(row + "\n")

    errors.close()
    # print("toCheck ", toCheck)
    print("totalCities= ", len(cityDataCities))
    print("totalPassed= ", totalPassed, " and totalFailed= ", totalFailed)

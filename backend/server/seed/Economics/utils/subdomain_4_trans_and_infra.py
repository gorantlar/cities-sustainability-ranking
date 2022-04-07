from itertools import count

import openpyxl
import csv




def parseFile(file):
    # read excel
    wrkbk = openpyxl.load_workbook(file)
    sh = wrkbk.active
    cities_map = {}

    # for row in sh.iter_rows(min_row=0, min_col=0, max_row=12, max_col=3):
    for row in sh.iter_rows(min_row=2, min_col=0):
        state_code = row[0].value.strip().lower()
        city_name = row[4].value.strip().lower()
        metric_name = row[5].value.strip().lower()
        val = row[11].value
        key = city_name + ":" + state_code
        if key not in cities_map:
            cities_map[key] = {}

        cities_map[key][metric_name] = val
    return cities_map

if __name__ == '__main__':
    # https://www2.census.gov/programs-surveys/economic-census/data/2017/sector00/EC1700BASIC.zip
    cities_details = parseFile("../excels/CHDB_data_city_all_2019.xlsx")
    # for county, val in counties_map.items():
    #     print(county, ": ", val)
    #     for key in val:
    #         temp = "{0:.2f}%".format(val[key][0]/val[key][1])
    #         print(f'{key}(%) = {temp}')

    # As details are per races
    t_and_i_columns = ["broadband_connection_2019", "park_access_2018", "preventive_services_2018", "walkability_2019"]
    metric_names = ["broadband connection", "park access", "preventive services, 65+", "walkability"]

    # states not present on city-data.com
    exceptions_states = {"puerto rico", "village of islands"}

    # name of csv output file
    output_file = open("../output/output.csv", 'w', newline='', encoding='utf-8')
    csv_writer = csv.writer(output_file)


    def writeCityDetails(row, city_details_local):
        rowLenOffset = len(row)
        print(city_details_local)
        # initialize all the values
        row += ["N"] * len(t_and_i_columns)
        if city_details_local:
            for ind, name in enumerate(metric_names):
                index = rowLenOffset + ind
                row[index] = city_details_local[name]
        print(row)
        csv_writer.writerow(row)


    # For Manual-review
    to_check = []
    total_passed = 0
    total_failed = 0

    with open("../final500Cities.csv", "r") as read_obj:
        csv_reader = csv.reader(read_obj)
        headers = next(csv_reader)
        baseHeadersLen = len(headers)


        # add new columns
        def add_new_columns():
            headers.extend(t_and_i_columns)


        add_new_columns()
        # writing the headers
        csv_writer.writerow(headers)
        # print("headers: ", headers)

        for row in csv_reader:
            state_name = row[2].lower().strip()
            state_code = row[3].lower().strip()
            city_name = row[1].lower().strip()
            key = city_name + ":" + state_code
            city_details = {}

            if key in cities_details:
                city_details = cities_details[key]
                total_passed += 1
            else:
                if state_name not in exceptions_states:
                    # closestMatchList = findClosestMatch(key)
                    # print("closest match for ", key, " are ", closestMatchList)
                    to_check.append([key, city_details])
                    total_failed += 1
            writeCityDetails(row, city_details)

    # todo
    # Target cities score (check with 500 cities)

    # print("totalCities= ", len(cityDataCities))
    print("totalPassed= ", total_passed, " and totalFailed= ", total_failed)
    print("toCheck: ", to_check)

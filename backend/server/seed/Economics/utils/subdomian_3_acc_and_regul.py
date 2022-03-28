import openpyxl
import csv


def parseFile(file):
    # read excel
    wrkbk = openpyxl.load_workbook(file)
    sh = wrkbk.active
    states_map = {}

    def get_parsed_percents(col_val):
        arr = col_val.strip().split()
        if len(arr) < 3:
            return -1
        # eg: Less than 10%
        if len(arr) < 4:
            return 5
        # eg: 30% to less than 40%
        start = int(arr[0].replace("%", ""))
        end = int(arr[-1].replace("%", ""))
        return (start + end) >> 1

    # fileName = file.split("\\")[-1]
    # stateName = " ".join(fileName.split()[:-1]).lower()

    # for row in sh.iter_rows(min_row=0, min_col=0, max_row=12, max_col=3):
    for row in sh.iter_rows(min_row=3, min_col=0):
        state_name = row[1].value.lower()
        year = row[10].value
        # Range indicating percent of total sales, value of shipments, or revenue imputed
        revenue_imputed = get_parsed_percents(row[17].value)
        # Range indicating percent of total annual payroll imputed
        payroll_imputed = get_parsed_percents(row[18].value)
        # Range indicating percent of total employees imputed
        employees_imputed = get_parsed_percents(row[19].value)
        if year != 2017:
            print(f'Year changed: {year}')
        if state_name not in states_map:
            states_map[state_name] = {}
            # will store cumulative sum of (avg %) and count of valid entries
            states_map[state_name]["revenue_imputed_2017"] = [0, 0]
            states_map[state_name]["payroll_imputed_2017"] = [0, 0]
            states_map[state_name]["employees_imputed_2017"] = [0, 0]

        if revenue_imputed != -1:
            states_map[state_name]["revenue_imputed_2017"][0] += revenue_imputed
            states_map[state_name]["revenue_imputed_2017"][1] += 1

        if payroll_imputed != -1:
            states_map[state_name]["payroll_imputed_2017"][0] += payroll_imputed
            states_map[state_name]["payroll_imputed_2017"][1] += 1

        if employees_imputed != -1:
            states_map[state_name]["employees_imputed_2017"][0] += employees_imputed
            states_map[state_name]["employees_imputed_2017"][1] += 1

    return states_map


if __name__ == '__main__':
    # https://www2.census.gov/programs-surveys/economic-census/data/2017/sector00/EC1700BASIC.zip
    states_map = parseFile("../excels/subdomain_3_A&R.xlsx")
    # for county, val in counties_map.items():
    #     print(county, ": ", val)
    #     for key in val:
    #         temp = "{0:.2f}%".format(val[key][0]/val[key][1])
    #         print(f'{key}(%) = {temp}')

    # As details are per races
    a_and_r_columns = ["revenue_imputed_2017", "payroll_imputed_2017", "employees_imputed_2017"]

    # states not present on city-data.com
    exceptions_states = {"puerto rico", "village of islands"}

    # name of csv output file
    output_file = open("../output/output.csv", 'w', newline='', encoding='utf-8')
    csv_writer = csv.writer(output_file)


    def writeCityDetails(row, city_details_local):
        rowLenOffset = len(row)

        # initialize all the values
        row += ["N"] * len(a_and_r_columns)
        if city_details_local:
            val_1 = city_details_local["revenue_imputed_2017"]
            temp = "{0:.2f}%".format(val_1[0] / val_1[1])
            row[rowLenOffset] = temp

            val_2 = city_details_local["payroll_imputed_2017"]
            temp = "{0:.2f}%".format(val_2[0] / val_2[1])
            row[rowLenOffset + 1] = temp

            val_3 = city_details_local["employees_imputed_2017"]
            temp = "{0:.2f}%".format(val_3[0] / val_3[1])
            row[rowLenOffset + 2] = temp
        csv_writer.writerow(row)


    # For Manual-review
    to_check = []
    total_passed = 0
    total_failed = 0

    with open("../final500Cities.csv", "r") as read_obj:
        csv_reader = csv.reader(read_obj)
        headers = next(csv_reader)
        baseHeadersLen = len(headers)


        # add new columns
        def add_new_columns():
            headers.extend(a_and_r_columns)


        add_new_columns()
        # writing the headers
        csvWriter.writerow(headers)
        # print("headers: ", headers)

        for row in csv_reader:
            state_name = row[2].lower()
            city_details = {}

            if state_name in states_map:
                city_details = states_map[state_name]
                total_passed += 1
            else:
                if state_name not in exceptions_states:
                    # closestMatchList = findClosestMatch(key)
                    # print("closest match for ", key, " are ", closestMatchList)
                    to_check.append([state_name, city_details])
                    total_failed += 1
            writeCityDetails(row, city_details)

    # todo
    # Target cities score (check with 500 cities)

    # print("totalCities= ", len(cityDataCities))
    print("totalPassed= ", total_passed, " and totalFailed= ", total_failed)
    print("toCheck: ", to_check)
